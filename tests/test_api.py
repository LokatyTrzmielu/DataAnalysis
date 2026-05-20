"""Integration tests for the FastAPI endpoints."""

from pathlib import Path

import pytest
import pytest_asyncio
from httpx import AsyncClient

FIXTURES_DIR = Path(__file__).parent / "fixtures"


# ============================================================================
# Auth
# ============================================================================


class TestAuth:
    async def test_create_user(self, client: AsyncClient):
        resp = await client.post(
            "/api/v1/auth/users",
            json={"email": "new@example.com", "name": "New User", "password": "pass1234"},
        )
        assert resp.status_code == 201
        data = resp.json()
        assert data["email"] == "new@example.com"
        assert data["name"] == "New User"
        assert "id" in data

    async def test_create_user_duplicate_email(self, client: AsyncClient):
        payload = {"email": "dup@example.com", "name": "User", "password": "pass1234"}
        await client.post("/api/v1/auth/users", json=payload)
        resp = await client.post("/api/v1/auth/users", json=payload)
        assert resp.status_code == 409

    async def test_login_success(self, client: AsyncClient):
        await client.post(
            "/api/v1/auth/users",
            json={"email": "login@example.com", "name": "Login User", "password": "secret"},
        )
        resp = await client.post(
            "/api/v1/auth/login",
            json={"email": "login@example.com", "password": "secret"},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert "access_token" in data
        assert data["email"] == "login@example.com"

    async def test_login_wrong_password(self, client: AsyncClient):
        await client.post(
            "/api/v1/auth/users",
            json={"email": "user@example.com", "name": "User", "password": "correct"},
        )
        resp = await client.post(
            "/api/v1/auth/login",
            json={"email": "user@example.com", "password": "wrong"},
        )
        assert resp.status_code == 401

    async def test_login_unknown_user(self, client: AsyncClient):
        resp = await client.post(
            "/api/v1/auth/login",
            json={"email": "nobody@example.com", "password": "any"},
        )
        assert resp.status_code == 401

    async def test_me(self, client: AsyncClient, auth_headers: dict):
        resp = await client.get("/api/v1/auth/me", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["email"] == "test@example.com"

    async def test_me_no_token(self, client: AsyncClient):
        resp = await client.get("/api/v1/auth/me")
        assert resp.status_code == 401

    async def test_me_invalid_token(self, client: AsyncClient):
        resp = await client.get("/api/v1/auth/me", headers={"Authorization": "Bearer invalid"})
        assert resp.status_code == 401

    async def test_change_password(self, client: AsyncClient, auth_headers: dict):
        resp = await client.post(
            "/api/v1/auth/change-password",
            json={"old_password": "testpass123", "new_password": "newpass456"},
            headers=auth_headers,
        )
        assert resp.status_code == 204

    async def test_change_password_wrong_old(self, client: AsyncClient, auth_headers: dict):
        resp = await client.post(
            "/api/v1/auth/change-password",
            json={"old_password": "wrong", "new_password": "new"},
            headers=auth_headers,
        )
        assert resp.status_code == 400

    async def test_list_users(self, client: AsyncClient, auth_headers: dict):
        # Create a second user
        await client.post(
            "/api/v1/auth/users",
            json={"email": "other@example.com", "name": "Other", "password": "pass"},
        )
        resp = await client.get("/api/v1/auth/users", headers=auth_headers)
        assert resp.status_code == 200
        emails = [u["email"] for u in resp.json()]
        assert "other@example.com" in emails
        assert "test@example.com" not in emails  # own user excluded


# ============================================================================
# Runs CRUD
# ============================================================================


class TestRunsCRUD:
    async def test_create_run(self, client: AsyncClient, auth_headers: dict):
        resp = await client.post(
            "/api/v1/runs",
            json={"client_name": "Test Client"},
            headers=auth_headers,
        )
        assert resp.status_code == 201
        data = resp.json()
        assert data["client_name"] == "Test Client"
        assert data["status"] == "created"
        assert data["quality_result"] is None

    async def test_list_runs_empty(self, client: AsyncClient, auth_headers: dict):
        resp = await client.get("/api/v1/runs", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 0
        assert data["items"] == []

    async def test_list_runs_with_data(self, client: AsyncClient, auth_headers: dict):
        for name in ["Alpha", "Beta", "Gamma"]:
            await client.post("/api/v1/runs", json={"client_name": name}, headers=auth_headers)

        resp = await client.get("/api/v1/runs", headers=auth_headers)
        assert resp.status_code == 200
        assert resp.json()["total"] == 3

    async def test_list_runs_search(self, client: AsyncClient, auth_headers: dict):
        for name in ["AlphaClient", "BetaClient", "GammaClient"]:
            await client.post("/api/v1/runs", json={"client_name": name}, headers=auth_headers)

        resp = await client.get("/api/v1/runs?search=Alpha", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 1
        assert data["items"][0]["client_name"] == "AlphaClient"

    async def test_list_runs_status_filter(self, client: AsyncClient, auth_headers: dict):
        await client.post("/api/v1/runs", json={"client_name": "C1"}, headers=auth_headers)

        resp = await client.get("/api/v1/runs?status_filter=created", headers=auth_headers)
        assert resp.status_code == 200
        assert resp.json()["total"] >= 1

        resp = await client.get("/api/v1/runs?status_filter=capacity_done", headers=auth_headers)
        assert resp.status_code == 200
        assert resp.json()["total"] == 0

    async def test_get_run(self, client: AsyncClient, auth_headers: dict):
        cr = await client.post("/api/v1/runs", json={"client_name": "X"}, headers=auth_headers)
        run_id = cr.json()["id"]

        resp = await client.get(f"/api/v1/runs/{run_id}", headers=auth_headers)
        assert resp.status_code == 200
        assert resp.json()["id"] == run_id

    async def test_get_run_not_found(self, client: AsyncClient, auth_headers: dict):
        resp = await client.get("/api/v1/runs/00000000-0000-0000-0000-000000000000", headers=auth_headers)
        assert resp.status_code == 404

    async def test_patch_run(self, client: AsyncClient, auth_headers: dict):
        cr = await client.post("/api/v1/runs", json={"client_name": "Original"}, headers=auth_headers)
        run_id = cr.json()["id"]

        resp = await client.patch(
            f"/api/v1/runs/{run_id}",
            json={"client_name": "Updated", "notes": "Some notes", "is_public": True},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["client_name"] == "Updated"
        assert data["notes"] == "Some notes"
        assert data["is_public"] is True

    async def test_delete_run(self, client: AsyncClient, auth_headers: dict):
        cr = await client.post("/api/v1/runs", json={"client_name": "ToDelete"}, headers=auth_headers)
        run_id = cr.json()["id"]

        resp = await client.delete(f"/api/v1/runs/{run_id}", headers=auth_headers)
        assert resp.status_code == 204

        resp = await client.get(f"/api/v1/runs/{run_id}", headers=auth_headers)
        assert resp.status_code == 404

    async def test_duplicate_run(self, client: AsyncClient, auth_headers: dict):
        cr = await client.post("/api/v1/runs", json={"client_name": "Original"}, headers=auth_headers)
        run_id = cr.json()["id"]

        resp = await client.post(f"/api/v1/runs/{run_id}/duplicate", headers=auth_headers)
        assert resp.status_code == 201
        data = resp.json()
        assert data["client_name"] == "Original (copy)"
        assert data["id"] != run_id
        assert data["status"] == "created"

    async def test_access_denied_for_private_run(self, client: AsyncClient, auth_headers: dict):
        # Create run as user1 (auth_headers)
        cr = await client.post("/api/v1/runs", json={"client_name": "Private"}, headers=auth_headers)
        run_id = cr.json()["id"]

        # Register and login as user2
        await client.post(
            "/api/v1/auth/users",
            json={"email": "user2@example.com", "name": "User2", "password": "pass2"},
        )
        resp2 = await client.post(
            "/api/v1/auth/login",
            json={"email": "user2@example.com", "password": "pass2"},
        )
        headers2 = {"Authorization": f"Bearer {resp2.json()['access_token']}"}

        resp = await client.get(f"/api/v1/runs/{run_id}", headers=headers2)
        assert resp.status_code == 403

    async def test_public_run_accessible_by_others(self, client: AsyncClient, auth_headers: dict):
        cr = await client.post("/api/v1/runs", json={"client_name": "Public"}, headers=auth_headers)
        run_id = cr.json()["id"]
        await client.patch(f"/api/v1/runs/{run_id}", json={"is_public": True}, headers=auth_headers)

        await client.post(
            "/api/v1/auth/users",
            json={"email": "viewer@example.com", "name": "Viewer", "password": "pass"},
        )
        resp_v = await client.post(
            "/api/v1/auth/login",
            json={"email": "viewer@example.com", "password": "pass"},
        )
        headers_v = {"Authorization": f"Bearer {resp_v.json()['access_token']}"}

        resp = await client.get(f"/api/v1/runs/{run_id}", headers=headers_v)
        assert resp.status_code == 200

    async def test_pagination(self, client: AsyncClient, auth_headers: dict):
        for i in range(5):
            await client.post("/api/v1/runs", json={"client_name": f"Run{i}"}, headers=auth_headers)

        resp = await client.get("/api/v1/runs?page=1&page_size=2", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 5
        assert len(data["items"]) == 2


# ============================================================================
# Analysis pipeline
# ============================================================================


class TestAnalysisPipeline:
    """Test the full analysis pipeline through the API."""

    async def _create_run(self, client: AsyncClient, headers: dict) -> str:
        resp = await client.post("/api/v1/runs", json={"client_name": "Pipeline Test"}, headers=headers)
        return resp.json()["id"]

    async def test_masterdata_inspect(self, client: AsyncClient, auth_headers: dict):
        run_id = await self._create_run(client, auth_headers)

        with open(FIXTURES_DIR / "masterdata_clean.csv", "rb") as f:
            resp = await client.post(
                f"/api/v1/runs/{run_id}/masterdata/inspect",
                files={"file": ("masterdata_clean.csv", f, "text/csv")},
                headers=auth_headers,
            )
        assert resp.status_code == 200
        data = resp.json()
        assert "file_columns" in data
        assert "suggestions" in data
        assert "preview_rows" in data
        assert len(data["preview_rows"]) > 0
        assert "sku" in data["file_columns"]

    async def test_quality_check(self, client: AsyncClient, auth_headers: dict):
        run_id = await self._create_run(client, auth_headers)

        with open(FIXTURES_DIR / "masterdata_clean.csv", "rb") as f:
            resp = await client.post(
                f"/api/v1/runs/{run_id}/quality",
                files={"file": ("masterdata_clean.csv", f, "text/csv")},
                headers=auth_headers,
            )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "quality_done"
        qr = data["quality_result"]
        assert qr is not None
        assert qr["total_records"] > 0
        assert 0 <= qr["dimensions_coverage_pct"] <= 100
        assert 0 <= qr["overall_score"] <= 100

    async def test_capacity_analysis(self, client: AsyncClient, auth_headers: dict):
        run_id = await self._create_run(client, auth_headers)

        # Upload masterdata via quality check
        with open(FIXTURES_DIR / "masterdata_clean.csv", "rb") as f:
            await client.post(
                f"/api/v1/runs/{run_id}/quality",
                files={"file": ("masterdata_clean.csv", f, "text/csv")},
                headers=auth_headers,
            )

        # Run capacity (uses saved masterdata_path)
        resp = await client.post(
            f"/api/v1/runs/{run_id}/capacity",
            data={"prioritization_mode": "false", "best_fit_mode": "false", "borderline_threshold": "2.0"},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "capacity_done"
        cr = data["capacity_result"]
        assert cr is not None
        assert cr["total_sku"] > 0
        assert 0 <= cr["fit_percentage"] <= 100
        assert len(cr["carriers_analyzed"]) > 0

    async def test_orders_inspect(self, client: AsyncClient, auth_headers: dict):
        run_id = await self._create_run(client, auth_headers)

        with open(FIXTURES_DIR / "orders_clean.csv", "rb") as f:
            resp = await client.post(
                f"/api/v1/runs/{run_id}/orders/inspect",
                files={"file": ("orders_clean.csv", f, "text/csv")},
                headers=auth_headers,
            )
        assert resp.status_code == 200
        data = resp.json()
        assert "file_columns" in data
        assert "sku" in data["file_columns"]
        assert len(data["suggestions"]) > 0

    async def test_orders_ingest(self, client: AsyncClient, auth_headers: dict):
        run_id = await self._create_run(client, auth_headers)

        with open(FIXTURES_DIR / "orders_clean.csv", "rb") as f:
            await client.post(
                f"/api/v1/runs/{run_id}/orders/inspect",
                files={"file": ("orders_clean.csv", f, "text/csv")},
                headers=auth_headers,
            )

        resp = await client.post(
            f"/api/v1/runs/{run_id}/orders/ingest",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "orders_ingested"
        assert data["analysis_config"]["orders_rows"] > 0

    async def test_performance_analysis(self, client: AsyncClient, auth_headers: dict):
        run_id = await self._create_run(client, auth_headers)

        # Upload and ingest orders
        with open(FIXTURES_DIR / "orders_clean.csv", "rb") as f:
            await client.post(
                f"/api/v1/runs/{run_id}/orders/inspect",
                files={"file": ("orders_clean.csv", f, "text/csv")},
                headers=auth_headers,
            )
        await client.post(f"/api/v1/runs/{run_id}/orders/ingest", headers=auth_headers)

        resp = await client.post(
            f"/api/v1/runs/{run_id}/performance",
            data={"productive_hours": "7.0"},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "performance_done"
        kpi = data["performance_result"]["kpi"]
        assert kpi["total_lines"] > 0
        assert kpi["total_orders"] > 0
        assert kpi["unique_sku"] > 0
        assert len(data["performance_result"]["sku_pareto"]) > 0

    async def test_capacity_without_masterdata_fails(self, client: AsyncClient, auth_headers: dict):
        run_id = await self._create_run(client, auth_headers)
        resp = await client.post(
            f"/api/v1/runs/{run_id}/capacity",
            data={"prioritization_mode": "false"},
            headers=auth_headers,
        )
        assert resp.status_code == 422

    async def test_performance_without_orders_fails(self, client: AsyncClient, auth_headers: dict):
        run_id = await self._create_run(client, auth_headers)
        resp = await client.post(
            f"/api/v1/runs/{run_id}/performance",
            data={"productive_hours": "7.0"},
            headers=auth_headers,
        )
        assert resp.status_code == 422

    async def test_orders_ingest_without_inspect_fails(self, client: AsyncClient, auth_headers: dict):
        run_id = await self._create_run(client, auth_headers)
        resp = await client.post(f"/api/v1/runs/{run_id}/orders/ingest", headers=auth_headers)
        assert resp.status_code == 422


# ============================================================================
# Reports
# ============================================================================


@pytest_asyncio.fixture
async def run_with_full_pipeline(client: AsyncClient, auth_headers: dict) -> str:
    """Create a run and execute the full analysis pipeline. Returns run_id."""
    cr = await client.post(
        "/api/v1/runs",
        json={"client_name": "ReportTest"},
        headers=auth_headers,
    )
    run_id = cr.json()["id"]

    # Quality
    with open(FIXTURES_DIR / "masterdata_clean.csv", "rb") as f:
        await client.post(
            f"/api/v1/runs/{run_id}/quality",
            files={"file": ("masterdata_clean.csv", f, "text/csv")},
            headers=auth_headers,
        )

    # Capacity
    await client.post(
        f"/api/v1/runs/{run_id}/capacity",
        data={"prioritization_mode": "false", "best_fit_mode": "false"},
        headers=auth_headers,
    )

    # Orders
    with open(FIXTURES_DIR / "orders_clean.csv", "rb") as f:
        await client.post(
            f"/api/v1/runs/{run_id}/orders/inspect",
            files={"file": ("orders_clean.csv", f, "text/csv")},
            headers=auth_headers,
        )
    await client.post(f"/api/v1/runs/{run_id}/orders/ingest", headers=auth_headers)
    await client.post(
        f"/api/v1/runs/{run_id}/performance",
        data={"productive_hours": "7.0"},
        headers=auth_headers,
    )

    return run_id


class TestReports:
    @pytest.mark.parametrize("report_name", [
        "DQ_Summary",
        "DQ_MissingCritical",
        "DQ_SuspectOutliers",
        "DQ_HighRiskBorderline",
        "DQ_Duplicates",
        "DQ_Conflicts",
    ])
    async def test_dq_xlsx_reports(
        self,
        client: AsyncClient,
        auth_headers: dict,
        run_with_full_pipeline: str,
        report_name: str,
    ):
        resp = await client.get(
            f"/api/v1/runs/{run_with_full_pipeline}/reports/xlsx/{report_name}",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        # xlsx files are zip archives starting with "PK"
        assert resp.content[:2] == b"PK"

    async def test_capacity_results_xlsx(
        self, client: AsyncClient, auth_headers: dict, run_with_full_pipeline: str
    ):
        import io
        from openpyxl import load_workbook

        resp = await client.get(
            f"/api/v1/runs/{run_with_full_pipeline}/reports/xlsx/Capacity_Results",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        # Primary regression guard: response is a valid xlsx workbook.
        wb = load_workbook(io.BytesIO(resp.content))
        assert "Capacity_Results" in wb.sheetnames

    async def test_sku_pareto_xlsx(
        self, client: AsyncClient, auth_headers: dict, run_with_full_pipeline: str
    ):
        import io
        from openpyxl import load_workbook

        resp = await client.get(
            f"/api/v1/runs/{run_with_full_pipeline}/reports/xlsx/SKU_Pareto",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "abc_class" in headers

    async def test_unknown_xlsx_report_returns_404(
        self, client: AsyncClient, auth_headers: dict, run_with_full_pipeline: str
    ):
        resp = await client.get(
            f"/api/v1/runs/{run_with_full_pipeline}/reports/xlsx/NonExistentReport",
            headers=auth_headers,
        )
        assert resp.status_code == 404

    async def test_zip_download(
        self, client: AsyncClient, auth_headers: dict, run_with_full_pipeline: str
    ):
        resp = await client.get(
            f"/api/v1/runs/{run_with_full_pipeline}/reports/zip",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/zip"
        # Verify it's a valid ZIP
        import io
        import zipfile
        zf = zipfile.ZipFile(io.BytesIO(resp.content))
        assert len(zf.namelist()) > 0

    async def test_pdf_download(
        self, client: AsyncClient, auth_headers: dict, run_with_full_pipeline: str
    ):
        resp = await client.get(
            f"/api/v1/runs/{run_with_full_pipeline}/reports/pdf",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "application/pdf" in resp.headers["content-type"]
        # Valid PDF starts with %PDF
        assert resp.content[:4] == b"%PDF"

    async def test_pdf_without_capacity_returns_422(
        self, client: AsyncClient, auth_headers: dict
    ):
        # Run without capacity analysis
        cr = await client.post(
            "/api/v1/runs", json={"client_name": "NoCap"}, headers=auth_headers
        )
        run_id = cr.json()["id"]
        resp = await client.get(f"/api/v1/runs/{run_id}/reports/pdf", headers=auth_headers)
        assert resp.status_code == 422

    async def test_xlsx_without_quality_returns_422(
        self, client: AsyncClient, auth_headers: dict
    ):
        cr = await client.post(
            "/api/v1/runs", json={"client_name": "NoQuality"}, headers=auth_headers
        )
        run_id = cr.json()["id"]
        resp = await client.get(
            f"/api/v1/runs/{run_id}/reports/xlsx/DQ_Summary", headers=auth_headers
        )
        assert resp.status_code == 422

    async def test_capacity_xlsx_without_capacity_returns_422(
        self, client: AsyncClient, auth_headers: dict
    ):
        cr = await client.post(
            "/api/v1/runs", json={"client_name": "NoCap2"}, headers=auth_headers
        )
        run_id = cr.json()["id"]
        resp = await client.get(
            f"/api/v1/runs/{run_id}/reports/xlsx/Capacity_Results", headers=auth_headers
        )
        assert resp.status_code == 422

    async def test_sku_pareto_without_performance_returns_422(
        self, client: AsyncClient, auth_headers: dict
    ):
        cr = await client.post(
            "/api/v1/runs", json={"client_name": "NoPerf"}, headers=auth_headers
        )
        run_id = cr.json()["id"]
        resp = await client.get(
            f"/api/v1/runs/{run_id}/reports/xlsx/SKU_Pareto", headers=auth_headers
        )
        assert resp.status_code == 422

    async def test_reports_access_denied_for_other_user(
        self, client: AsyncClient, auth_headers: dict, run_with_full_pipeline: str
    ):
        await client.post(
            "/api/v1/auth/users",
            json={"email": "outsider@example.com", "name": "Out", "password": "pass"},
        )
        resp_o = await client.post(
            "/api/v1/auth/login",
            json={"email": "outsider@example.com", "password": "pass"},
        )
        headers_o = {"Authorization": f"Bearer {resp_o.json()['access_token']}"}

        resp = await client.get(
            f"/api/v1/runs/{run_with_full_pipeline}/reports/xlsx/DQ_Summary",
            headers=headers_o,
        )
        assert resp.status_code == 403


# ============================================================================
# Sharing
# ============================================================================


class TestSharing:
    async def _create_second_user(self, client: AsyncClient) -> dict:
        await client.post(
            "/api/v1/auth/users",
            json={"email": "share_target@example.com", "name": "Share Target", "password": "pass"},
        )
        resp = await client.post(
            "/api/v1/auth/login",
            json={"email": "share_target@example.com", "password": "pass"},
        )
        return {"Authorization": f"Bearer {resp.json()['access_token']}"}

    async def test_share_run_with_user(self, client: AsyncClient, auth_headers: dict):
        await client.post(
            "/api/v1/auth/users",
            json={"email": "share_target@example.com", "name": "Target", "password": "pass"},
        )
        cr = await client.post("/api/v1/runs", json={"client_name": "Shared"}, headers=auth_headers)
        run_id = cr.json()["id"]

        resp = await client.post(
            f"/api/v1/runs/{run_id}/shares",
            json={"email": "share_target@example.com"},
            headers=auth_headers,
        )
        assert resp.status_code == 201
        assert resp.json()["email"] == "share_target@example.com"

    async def test_shared_run_accessible(self, client: AsyncClient, auth_headers: dict):
        headers2 = await self._create_second_user(client)
        cr = await client.post("/api/v1/runs", json={"client_name": "SharedRun"}, headers=auth_headers)
        run_id = cr.json()["id"]
        await client.post(
            f"/api/v1/runs/{run_id}/shares",
            json={"email": "share_target@example.com"},
            headers=auth_headers,
        )
        resp = await client.get(f"/api/v1/runs/{run_id}", headers=headers2)
        assert resp.status_code == 200

    async def test_list_shares(self, client: AsyncClient, auth_headers: dict):
        await client.post(
            "/api/v1/auth/users",
            json={"email": "share_list@example.com", "name": "ShareList", "password": "pass"},
        )
        cr = await client.post("/api/v1/runs", json={"client_name": "X"}, headers=auth_headers)
        run_id = cr.json()["id"]
        await client.post(
            f"/api/v1/runs/{run_id}/shares",
            json={"email": "share_list@example.com"},
            headers=auth_headers,
        )
        resp = await client.get(f"/api/v1/runs/{run_id}/shares", headers=auth_headers)
        assert resp.status_code == 200
        assert len(resp.json()) == 1

    async def test_remove_share(self, client: AsyncClient, auth_headers: dict):
        await client.post(
            "/api/v1/auth/users",
            json={"email": "rm_share@example.com", "name": "RmShare", "password": "pass"},
        )
        cr = await client.post("/api/v1/runs", json={"client_name": "Y"}, headers=auth_headers)
        run_id = cr.json()["id"]
        share_resp = await client.post(
            f"/api/v1/runs/{run_id}/shares",
            json={"email": "rm_share@example.com"},
            headers=auth_headers,
        )
        user_id = share_resp.json()["user_id"]

        resp = await client.delete(f"/api/v1/runs/{run_id}/shares/{user_id}", headers=auth_headers)
        assert resp.status_code == 204

    async def test_share_with_nonexistent_user(self, client: AsyncClient, auth_headers: dict):
        cr = await client.post("/api/v1/runs", json={"client_name": "Z"}, headers=auth_headers)
        run_id = cr.json()["id"]
        resp = await client.post(
            f"/api/v1/runs/{run_id}/shares",
            json={"email": "nobody@example.com"},
            headers=auth_headers,
        )
        assert resp.status_code == 404

    async def test_share_duplicate_rejected(self, client: AsyncClient, auth_headers: dict):
        await client.post(
            "/api/v1/auth/users",
            json={"email": "dup_share@example.com", "name": "DupShare", "password": "pass"},
        )
        cr = await client.post("/api/v1/runs", json={"client_name": "W"}, headers=auth_headers)
        run_id = cr.json()["id"]
        for _ in range(2):
            resp = await client.post(
                f"/api/v1/runs/{run_id}/shares",
                json={"email": "dup_share@example.com"},
                headers=auth_headers,
            )
        assert resp.status_code == 409


# ============================================================================
# Health check
# ============================================================================


async def test_health(client: AsyncClient):
    resp = await client.get("/health")
    assert resp.status_code == 200
    assert resp.json() == {"status": "ok"}
