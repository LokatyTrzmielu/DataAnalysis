"""Excel (xlsx) generator for the Container Order Calculator tool."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def _style_header(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
    ws.row_dimensions[row].height = 22
    # Use a string reference for freeze_panes. The previous form
    # `ws.cell(row=row+1, column=1)` invoked .cell(), which *creates* the cell
    # at A(row+1) as a side effect; that extends ws.max_row, so the next
    # ws.append() jumps a row, leaving row 2 empty and breaking Excel's
    # "sort by header" auto-detection.
    ws.freeze_panes = f"A{row + 1}"


def _autosize(ws) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = cell.value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def _sheet_order_summary(wb: Workbook, plan: Any) -> None:
    ws = wb.create_sheet("Order Summary", 0)
    # "Bases", "Frames" and "Dividers" make the procurement breakdown explicit.
    # Each physical bin needs 1 base (138 mm) + N frames (50 mm each) +
    # locations_per_bin divider cells.
    headers = ["Variant", "Footprint", "Bin height (mm)", "Cell L×W×H (mm)",
               "Locations / bin", "SKU count", "Locations total", "Bins to order",
               "Bases", "Frames", "Dividers", "Total weight (kg)",
               "Avg cell fill (%)"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    total_dividers = 0
    total_weight = 0.0
    for s in plan.summaries:
        cell = f"{s.cell_length_mm}×{s.cell_width_mm}×{s.cell_height_mm}"
        total_dividers += s.dividers_required
        total_weight += s.total_weight_kg
        ws.append([
            s.code, s.footprint_label, s.bin_height_mm, cell,
            s.locations_per_bin, s.sku_count, s.total_locations,
            s.bins_required,
            s.bins_required,            # Bases — 1 base per bin
            s.total_frames_required,    # Frames — bins × frames_per_bin
            s.dividers_required,        # Dividers — bins × locations_per_bin
            round(s.total_weight_kg, 2),
            s.avg_fill_pct,
        ])

    # TOTAL row.
    total_locations = sum(s.total_locations for s in plan.summaries)
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=6, value=plan.total_sku_covered)
    ws.cell(row=total_row, column=7, value=total_locations)
    ws.cell(row=total_row, column=8, value=plan.total_bins)
    ws.cell(row=total_row, column=9, value=plan.total_bins)        # Bases total
    ws.cell(row=total_row, column=10, value=plan.total_frames)     # Frames total
    ws.cell(row=total_row, column=11, value=total_dividers)        # Dividers total
    ws.cell(row=total_row, column=12, value=round(total_weight, 2))
    ws.cell(row=total_row, column=13, value=plan.avg_fill_pct)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=total_row, column=col)
        c.fill = TOTAL_FILL
        c.font = TOTAL_FONT

    _autosize(ws)


def _sheet_sku_assignment(wb: Workbook, plan: Any) -> None:
    ws = wb.create_sheet("SKU Assignment")
    # "Pcs/location" sits right after Variant — the headline number the
    # warehouse cares about: how many physical pieces of this SKU end up in
    # each allocated cell of the assigned variant.
    headers = ["SKU", "Fit", "Imputed", "Variant", "Pcs/location",
               "ABC", "Recommendation",
               "Length (mm)", "Width (mm)", "Height (mm)", "Weight (kg)",
               "Stock vol (L)", "Locations", "Bins", "Cell fill (%)"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for a in plan.assignments:
        pcs_loc = getattr(a, "pcs_per_location", 0) or 0
        ws.append([
            a.sku,
            a.fit_status or "",
            "Yes" if a.dimensions_imputed else "",
            a.variant_code or "—",
            pcs_loc if pcs_loc > 0 else "",
            a.abc_class or "",
            a.recommendation or "",
            a.length_mm, a.width_mm, a.height_mm, a.weight_kg, a.stock_volume_L,
            a.locations, a.bins, a.cell_fill_pct,
        ])
    _autosize(ws)


def _sheet_parameters(wb: Workbook, params: Any, run: Any, plan: Any,
                       user: Any = None) -> None:
    ws = wb.create_sheet("Parameters")
    ws.append(["Parameter", "Value"])
    _style_header(ws, 1, 2)

    generated_by = "—"
    if user is not None:
        name = getattr(user, "name", None) or ""
        email = getattr(user, "email", None) or ""
        generated_by = f"{name} <{email}>" if name and email else (name or email or "—")

    pairs = [
        ("Generated by", generated_by),
        ("Client", run.client_name),
        ("Generated at", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Mode", params.mode),
        ("Auto: goal", params.auto_goal if params.mode == "auto" else "—"),
        ("Auto: max variants", params.auto_max_variants if params.mode == "auto" else "—"),
        ("Guided: preset", params.guided_preset if params.mode == "guided" else "—"),
        ("Manual: variant codes", ", ".join(params.manual_variant_codes) if params.mode == "manual" else "—"),
        ("ABC classes", ", ".join(params.abc_classes) if params.abc_classes else "all"),
        ("Only Machine SKUs", "Yes" if params.only_machine else "No"),
        ("Include BORDERLINE", "Yes" if params.include_borderline else "No"),
        ("Stock multiplier", f"{params.stock_multiplier:.2f}"),
        ("Location fill rate", f"{params.location_fill_rate * 100:.0f}%"),
        ("Min locations / SKU", params.min_locations_per_sku),
        ("Max locations / SKU", params.max_locations_per_sku),
        ("", ""),
        ("Total bins to order", plan.total_bins),
        ("SKUs covered", f"{plan.total_sku_covered} of {plan.total_sku_planned} ({plan.coverage_pct:.1f}%)"),
        ("Average cell fill", f"{plan.avg_fill_pct:.1f}%"),
    ]
    for k, v in pairs:
        ws.append([k, v])
    _autosize(ws)


def _sheet_orphans(wb: Workbook, plan: Any) -> None:
    ws = wb.create_sheet("Orphans")
    headers = ["SKU", "Fit", "ABC", "Recommendation", "Length (mm)", "Width (mm)",
               "Height (mm)", "Weight (kg)", "Stock vol (L)", "Reason"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for a in plan.orphans:
        reason = (
            "Missing dimensions in source data"
            if a.orphan_reason == "missing_dimensions"
            else "No variant in selection fits this SKU"
        )
        ws.append([
            a.sku,
            a.fit_status or "",
            a.abc_class or "",
            a.recommendation or "",
            a.length_mm, a.width_mm, a.height_mm, a.weight_kg, a.stock_volume_L,
            reason,
        ])
    if not plan.orphans:
        ws.append(["(no orphan SKUs)", "", "", "", "", "", "", "", "", ""])
    _autosize(ws)


def generate_order_xlsx(plan: Any, params: Any, run: Any, user: Any = None) -> bytes:
    """Build the 4-sheet workbook.

    Args:
        plan: ContainerPlanResponse
        params: PlanParamsRequest
        run: SQLAlchemy AnalysisRun
        user: SQLAlchemy User (the person downloading the export). Used to fill
              the "Generated by" row in the Parameters sheet.
    """
    wb = Workbook()
    # openpyxl creates a default sheet — remove it (we'll add named sheets ourselves).
    default = wb.active
    if default is not None and default.title == "Sheet":
        wb.remove(default)

    _sheet_order_summary(wb, plan)
    _sheet_sku_assignment(wb, plan)
    _sheet_parameters(wb, params, run, plan, user)
    _sheet_orphans(wb, plan)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
