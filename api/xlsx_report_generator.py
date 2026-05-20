"""Single-sheet XLSX report generator used by the per-report download endpoint.

Replaces the previous CSV exports. Writing native xlsx avoids Excel's
locale-driven CSV interpretation — values like ``10.5`` (weight) were being
parsed as dates ("10 May") when CSVs were opened in Polish-locale Excel.
Numeric Python values flow straight into Excel cells as numbers here.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
CENTER = Alignment(horizontal="center", vertical="center")

# Column → Excel number format. Applied when the column appears in a report.
# Floats default to two decimals; integers stay un-formatted (General).
_NUM_FMT_FLOAT = "0.00"
_NUM_FMT_PCT_TEXT = "0.00"  # we already store these as plain numbers, not 0–1 fractions

NUMERIC_FORMATS: dict[str, str] = {
    # Capacity results — the columns that were misread as dates in CSV.
    "weight_kg": _NUM_FMT_FLOAT,
    "length_mm": _NUM_FMT_FLOAT,
    "width_mm": _NUM_FMT_FLOAT,
    "height_mm": _NUM_FMT_FLOAT,
    "volume_m3": "0.0000",
    "stock_volume_m3": "0.0000",
    "stored_volume_L": _NUM_FMT_FLOAT,
    "carrier_volume_L": _NUM_FMT_FLOAT,
    "filling_rate": _NUM_FMT_FLOAT,
    "margin_mm": _NUM_FMT_FLOAT,
    # DQ summary / Orders validation share simple percentage-as-number columns.
    "overall_score": _NUM_FMT_PCT_TEXT,
    "dimensions_coverage_pct": _NUM_FMT_PCT_TEXT,
    "weight_coverage_pct": _NUM_FMT_PCT_TEXT,
    "stock_coverage_pct": _NUM_FMT_PCT_TEXT,
    # Performance / Pareto
    "cumulative_pct": _NUM_FMT_PCT_TEXT,
    "cumulated_sku_pct": _NUM_FMT_PCT_TEXT,
    "lines_day_pct": _NUM_FMT_PCT_TEXT,
    "cumulated_lines_pct": _NUM_FMT_PCT_TEXT,
    "pieces_day_pct": _NUM_FMT_PCT_TEXT,
    "cumulated_pieces_pct": _NUM_FMT_PCT_TEXT,
}


def _style_header(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = f"A{row + 1}"


def _autosize(ws) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = cell.value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def _resolve_columns(rows: list[dict], columns: list[str] | None) -> list[str]:
    """Use explicit columns when provided; otherwise pick the first row's keys."""
    if columns:
        return columns
    if rows:
        return list(rows[0].keys())
    return []


def generate_report_xlsx(
    report_name: str,
    rows: list[dict],
    columns: list[str] | None = None,
) -> bytes:
    """Write a single-sheet workbook for ``rows`` and return the xlsx bytes.

    Numeric values are written natively (int/float) so Excel never re-interprets
    them as dates. Per-column number formats come from :data:`NUMERIC_FORMATS`.
    Empty result sets still produce a header-only sheet.
    """
    cols = _resolve_columns(rows, columns)

    wb = Workbook()
    ws = wb.active
    # Excel sheet names are capped at 31 chars and can't contain []:*?/\
    ws.title = (report_name[:31] or "Report").replace("/", "_").replace("\\", "_")

    if not cols:
        # Truly empty — nothing to write, return an empty workbook.
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    ws.append(cols)
    _style_header(ws, 1, len(cols))

    for r in rows:
        ws.append([_coerce(r.get(c)) for c in cols])

    # Apply numeric formats to known columns (data rows start at row 2).
    last_row = ws.max_row
    for col_idx, col_name in enumerate(cols, start=1):
        fmt = NUMERIC_FORMATS.get(col_name)
        if not fmt:
            continue
        for row_idx in range(2, last_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = fmt

    _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coerce(value: Any) -> Any:
    """Pass-through with one safety net: turn ``None`` into empty string only
    for cells where openpyxl would otherwise render the literal ``None``.

    openpyxl handles ``None`` correctly (empty cell), so we just pass through.
    """
    return value
